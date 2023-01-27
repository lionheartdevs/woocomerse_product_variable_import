#this imports openpyxl for working with excel speadsheets - dont remove
import openpyxl
#this uses the path of the file for the excel/csv file.
path = "D:\Example\Project_Folder\woocommerse_import_file.xlsx"
#this loads the excel sheets into memory-dont remove
wb = openpyxl.load_workbook(path)
#the active excel sheet is used
sheet = wb.active
##SETUP VARIABLES### - dont remove
count = 0 # for loop counter
row = 0 # what row the parent product is on
product_name = ""
attribute1 = 0 # custom metadata
attribute2 = 0 # custom metadata
attribute3 = 0 # custom metadata
attributeEmpty = False # on our list - we have some products with 2 variables and some with 3 - so this helped to alter the for loop.
parentskuid = "" # for setting the parentsku id for product variables.
###FOR LOOP STARTS####
#range will be the starting row number and the last row number + 1
for i in range(2, 3049 + 1):
    if count == 0:
        row = i
        cell = sheet.cell(row=i, column=1)
        cell2 = sheet.cell(row=i, column=2)
        cellvalue= str(cell2.value)
        if cellvalue.isnumeric():
            product_name = str(cell.value) + " " + cellvalue + "mg"
        else:
            product_name = str(cell.value) + " " + cellvalue
        cell.value = product_name
        parentskuid = "skuid" + str(i+i) + str(i)
        sheet.cell(row=i, column=22).value = parentskuid
        attribute1 = sheet.cell(row=i, column=3).value
        attribute2 = sheet.cell(row=i, column=6).value
        attribute3 = sheet.cell(row=i, column=9).value
        sheet.cell(row=i, column=17).value = "variable"
        sheet.cell(row=i, column=18).value = "default"
        sheet.cell(row=i, column=20).value = 1
        sheet.cell(row=i, column=21).value = "https://examplewebsite.com/image.jpg" # this is to set an image for the main product.
        if attribute3 == None:
            attributeEmpty = True
            sheet.cell(row=i, column=13).value = str(attribute1) + "," + str(attribute2)
            sheet.cell(row=i, column=14).value = sheet.cell(row=i, column=5).value
            sheet.cell(row=i, column=15).value = sheet.cell(row=i, column=8).value
            sheet.insert_rows(i + 1)
            sheet.cell(row=i + 1, column=17).value = "variation"
            sheet.cell(row=i + 1, column=19).value = "parent"
            sheet.cell(row=i + 1, column=22).value = parentskuid + "-1"
            sheet.cell(row=i + 1, column=23).value = parentskuid
            sheet.insert_rows(i + 2)
            sheet.cell(row=i + 2, column=17).value = "variation"
            sheet.cell(row=i + 2, column=19).value = "parent"
            sheet.cell(row=i + 2, column=22).value = parentskuid + "-2"
            sheet.cell(row=i + 2, column=23).value = parentskuid
            count += 1
        else:
            attributeEmpty = False
            sheet.cell(row=i, column=13).value = str(attribute1) + "," + str(attribute2) + "," + str(attribute3)
            sheet.cell(row=i, column=14).value = sheet.cell(row=i, column=5).value
            sheet.cell(row=i, column=15).value = sheet.cell(row=i, column=8).value
            sheet.cell(row=i, column=16).value = sheet.cell(row=i, column=11).value
            sheet.insert_rows(i + 1)
            sheet.cell(row=i + 1, column=17).value = "variation"
            sheet.cell(row=i + 1, column=19).value = "parent"
            sheet.cell(row=i + 1, column=22).value = parentskuid + "-1"
            sheet.cell(row=i + 1, column=23).value = parentskuid
            sheet.insert_rows(i + 2)
            sheet.cell(row=i + 2, column=17).value = "variation"
            sheet.cell(row=i + 2, column=19).value = "parent"
            sheet.cell(row=i + 2, column=22).value = parentskuid + "-2"
            sheet.cell(row=i + 2, column=23).value = parentskuid
            sheet.insert_rows(i + 3)
            sheet.cell(row=i + 3, column=17).value = "variation"
            sheet.cell(row=i + 3, column=19).value = "parent"
            sheet.cell(row=i + 3, column=22).value = parentskuid + "-3"
            sheet.cell(row=i + 3, column=23).value = parentskuid
            count += 1
    elif count == 1:
        cell = sheet.cell(row=i, column=1)
        cell.value = product_name + " " + "Tier 1"
        price = sheet.cell(row=row, column=5)
        newprice = sheet.cell(row=i, column=12)
        newprice.value = price.value
        sheet.cell(row=i, column=13).value = str(attribute1)
        count += 1
    elif count == 2:
        cell = sheet.cell(row=i, column=1)
        cell.value = product_name + " " + "Tier 2"
        price = sheet.cell(row=row, column=8)
        newprice = sheet.cell(row=i, column=12)
        newprice.value = price.value
        sheet.cell(row=i, column=13).value = str(attribute2)
        if attributeEmpty == True:
            count = 0
        else:
            count += 1
    elif count == 3:
        cell = sheet.cell(row=i, column=1)
        cell.value = product_name + " " + "Tier 3"
        price = sheet.cell(row=row, column=11)
        newprice = sheet.cell(row=i, column=12)
        newprice.value = price.value
        sheet.cell(row=i, column=13).value = str(attribute3)
        count = 0

wb.save(path)
